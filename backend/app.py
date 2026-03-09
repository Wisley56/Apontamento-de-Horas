"""
API Backend - Sistema de Apontamento de Horas
FastAPI server com endpoints para análise, exportação e persistência de histórico.
"""

from fastapi import FastAPI, HTTPException, Response, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Dict, Optional
from datetime import datetime, date
import os
import io
import sqlite3
import json
from pathlib import Path

from services.hours_service import process_day, time_to_decimal
from services.holidays_service import (
    get_holidays_for_period, 
    classify_date, 
    generate_date_range,
    get_states_list
)

# Importação para exportação Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = FastAPI(
    title="Sistema de Apontamento de Horas",
    description="API para controle e análise de apontamento de horas trabalhadas",
    version="2.0.0"
)

# CORS para permitir requisições do frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============ BANCO DE DADOS ============

# O arquivo .db fica na mesma pasta do backend
DB_PATH = Path(__file__).parent / "apontamentos.db"


def get_db():
    """Retorna uma conexão com o banco SQLite."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Inicializa o banco de dados criando a tabela se não existir."""
    conn = get_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS apontamentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                colaborador TEXT NOT NULL,
                periodo_inicio TEXT NOT NULL,
                periodo_fim TEXT NOT NULL,
                total_horas REAL DEFAULT 0,
                criado_em TEXT NOT NULL,
                dados_json TEXT NOT NULL
            )
        """)
        conn.commit()
    finally:
        conn.close()


@app.on_event("startup")
def on_startup():
    init_db()


# ============ MODELOS ============

class DateRange(BaseModel):
    start_date: str  # dd/mm/yyyy
    end_date: str    # dd/mm/yyyy


class ManualException(BaseModel):
    date: str        # dd/mm/yyyy
    type: str        # ferias, afastamento, atestado, banco, feriado_manual


class AnalyzeRequest(BaseModel):
    selection_type: str  # "period" ou "single"
    start_date: str      # dd/mm/yyyy
    end_date: Optional[str] = None  # dd/mm/yyyy (opcional se single)
    state: str           # UF
    worked_hours: List[str]  # Lista de HH:MM para cada dia
    manual_exceptions: Optional[List[ManualException]] = []


class DayResult(BaseModel):
    date: str
    day_of_week: str
    worked_time: str
    redmine_value: str
    difference: str
    status: str
    status_icon: str
    status_description: str
    css_class: str
    is_ignored: bool
    day_type: Optional[str]
    manual_status: Optional[str]


class AnalyzeResponse(BaseModel):
    days: List[DayResult]
    summary: Dict


class ExportRequest(BaseModel):
    days: List[Dict]


class IntervalDetail(BaseModel):
    entry: str = ""
    exit: str = ""
    description: str = ""


class DayDetail(BaseModel):
    date: str           # dd/mm/yyyy
    day_name: str
    intervals: List[IntervalDetail]
    total_hours: float = 0.0
    is_ignored: bool = False
    ignore_reason: str = ""


class SaveRequest(BaseModel):
    colaborador: str
    periodo_inicio: str   # dd/mm/yyyy
    periodo_fim: str      # dd/mm/yyyy
    total_horas: float = 0.0
    dias: List[DayDetail]


# ============ ENDPOINTS ============
# NOTA: A raiz "/" é servida automaticamente pelo StaticFiles(html=True) no final do arquivo


@app.get("/api/states")
async def list_states():
    """Lista todos os estados brasileiros"""
    return get_states_list()


@app.get("/api/holidays/{year}/{state}")
async def list_holidays(year: int, state: str):
    """Lista feriados de um ano para um estado"""
    try:
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        holidays_dict = get_holidays_for_period(start, end, state)
        return {"holidays": holidays_dict}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/analyze")
async def analyze_hours(request: AnalyzeRequest):
    """
    Analisa as horas trabalhadas para o período informado.
    """
    try:
        # Parse das datas
        start_date = datetime.strptime(request.start_date, "%d/%m/%Y").date()
        
        if request.selection_type == "single":
            end_date = start_date
        else:
            end_date = datetime.strptime(request.end_date, "%d/%m/%Y").date()
        
        if end_date < start_date:
            raise HTTPException(status_code=400, detail="Data final não pode ser menor que a inicial")
        
        # Gerar lista de datas
        dates = generate_date_range(start_date, end_date)
        
        # Verificar quantidade de horas informadas
        if len(request.worked_hours) != len(dates):
            raise HTTPException(
                status_code=400, 
                detail=f"Esperado {len(dates)} valores de horas, recebido {len(request.worked_hours)}"
            )
        
        # Buscar feriados
        holidays_dict = get_holidays_for_period(start_date, end_date, request.state)
        
        # Converter exceções manuais para dict
        manual_exceptions = {}
        if request.manual_exceptions:
            type_names = {
                "ferias": "Férias",
                "afastamento": "Afastamento",
                "atestado": "Atestado",
                "banco": "Banco de Horas",
                "feriado_manual": "Feriado Manual"
            }
            for exc in request.manual_exceptions:
                manual_exceptions[exc.date] = type_names.get(exc.type, exc.type)
        
        # Processar cada dia
        results = []
        stats = {
            "total_days": len(dates),
            "workdays_analyzed": 0,
            "days_ok": 0,
            "days_divergent": 0,
            "days_ignored": 0,
            "total_worked_hours": 0.0,
            "total_redmine_hours": 0.0,
            "conformity_percentage": 0.0
        }
        
        days_of_week = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
        
        for i, date_str in enumerate(dates):
            date_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
            day_of_week = days_of_week[date_obj.weekday()]
            
            # Classificar o dia
            classification = classify_date(date_str, holidays_dict, manual_exceptions)
            
            if classification["is_workday"]:
                # Dia útil - processar
                worked_time = request.worked_hours[i]
                day_result = process_day(date_str, worked_time)
                
                stats["workdays_analyzed"] += 1
                
                if day_result["status"] == "confere":
                    stats["days_ok"] += 1
                else:
                    stats["days_divergent"] += 1
                
                # Somar horas
                decimal_hours = time_to_decimal(worked_time)
                stats["total_worked_hours"] += decimal_hours
                stats["total_redmine_hours"] += decimal_hours
                
            else:
                # Dia ignorado
                day_result = process_day(date_str, "----", classification["description"])
                stats["days_ignored"] += 1
            
            day_result["day_of_week"] = day_of_week
            results.append(day_result)
        
        # Calcular percentual de conformidade
        if stats["workdays_analyzed"] > 0:
            stats["conformity_percentage"] = round(
                (stats["days_ok"] / stats["workdays_analyzed"]) * 100, 2
            )
        
        # Formatar totais
        stats["total_worked_display"] = f"{stats['total_worked_hours']:.2f}h"
        stats["total_redmine_display"] = f"{stats['total_redmine_hours']:.2f}h"
        
        return {
            "days": results,
            "summary": stats
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar datas: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export")
async def export_to_excel(request: ExportRequest):
    """
    Exporta os dados para arquivo Excel.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Apontamento de Horas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        divergent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ignored_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos (removido "Diferença")
        headers = ["Data", "Dia", "Tempo Trabalhado", "Valor Redmine", "Tipo", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dados
        for row_idx, day in enumerate(request.days, 2):
            ws.cell(row=row_idx, column=1, value=day.get("date", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=day.get("day_of_week", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=day.get("worked_time", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=day.get("redmine_value", "")).border = thin_border
            ws.cell(row=row_idx, column=5, value=day.get("day_type", "")).border = thin_border
            
            status_cell = ws.cell(row=row_idx, column=6, value=day.get("status_description", ""))
            status_cell.border = thin_border
            
            # Aplicar cor baseada no status
            status = day.get("status", "")
            if status == "ok":
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = ok_fill
            elif status == "divergent":
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = divergent_fill
            elif status == "ignorado":
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = ignored_fill
        
        # Ajustar largura das colunas
        column_widths = [12, 12, 18, 15, 20, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=apontamento_resultado.xlsx"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ ENDPOINTS DE HISTÓRICO ============

@app.get("/api/verificar-duplicata")
async def verificar_duplicata(
    colaborador: str = Query(..., description="Nome do colaborador"),
    periodo_inicio: str = Query(..., description="Data início dd/mm/yyyy"),
    periodo_fim: str = Query(..., description="Data fim dd/mm/yyyy"),
):
    """
    Verifica se já existe um apontamento para o mesmo colaborador que sobreponha
    o período informado. Retorna os registros conflitantes caso existam.
    """
    try:
        conn = get_db()
        try:
            # Busca registros do mesmo colaborador (case-insensitive)
            rows = conn.execute(
                """
                SELECT id, colaborador, periodo_inicio, periodo_fim, total_horas, criado_em
                FROM apontamentos
                WHERE LOWER(colaborador) = LOWER(?)
                ORDER BY id DESC
                """,
                (colaborador.strip(),)
            ).fetchall()
        finally:
            conn.close()

        if not rows:
            return {"duplicata": False, "registros": []}

        # Converter datas para comparação
        def parse_date(d):
            try:
                return datetime.strptime(d, "%d/%m/%Y").date()
            except Exception:
                return None

        novo_inicio = parse_date(periodo_inicio)
        novo_fim = parse_date(periodo_fim)

        if not novo_inicio or not novo_fim:
            return {"duplicata": False, "registros": []}

        conflitos = []
        for row in rows:
            ex_inicio = parse_date(row["periodo_inicio"])
            ex_fim = parse_date(row["periodo_fim"])
            if not ex_inicio or not ex_fim:
                continue
            # Sobreposição: não há sobreposição apenas se um termina antes do outro começar
            if not (novo_fim < ex_inicio or novo_inicio > ex_fim):
                conflitos.append({
                    "id": row["id"],
                    "colaborador": row["colaborador"],
                    "periodo_inicio": row["periodo_inicio"],
                    "periodo_fim": row["periodo_fim"],
                    "total_horas": row["total_horas"],
                    "criado_em": row["criado_em"],
                })

        return {
            "duplicata": len(conflitos) > 0,
            "registros": conflitos,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao verificar duplicata: {str(e)}")


@app.post("/api/salvar-apontamento", status_code=201)
async def salvar_apontamento(request: SaveRequest):
    """
    Salva um apontamento no banco de dados SQLite.
    Recebe: colaborador, periodo_inicio, periodo_fim, total_horas, dias (com intervalos e descrições).
    """
    try:
        criado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
        dados_json = json.dumps(
            [d.model_dump() for d in request.dias],
            ensure_ascii=False
        )

        conn = get_db()
        try:
            cursor = conn.execute(
                """
                INSERT INTO apontamentos (colaborador, periodo_inicio, periodo_fim, total_horas, criado_em, dados_json)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    request.colaborador.strip(),
                    request.periodo_inicio,
                    request.periodo_fim,
                    request.total_horas,
                    criado_em,
                    dados_json,
                )
            )
            conn.commit()
            record_id = cursor.lastrowid
        finally:
            conn.close()

        return {
            "id": record_id,
            "mensagem": "Apontamento salvo com sucesso!",
            "colaborador": request.colaborador,
            "criado_em": criado_em,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar apontamento: {str(e)}")


@app.get("/api/historico")
async def get_historico(
    colaborador: Optional[str] = Query(None, description="Filtrar por nome do colaborador"),
    mes: Optional[str] = Query(None, description="Filtrar por mês no formato YYYY-MM")
):
    """
    Retorna o histórico de apontamentos.
    Parâmetros opcionais: colaborador (texto), mes (YYYY-MM).
    """
    try:
        conn = get_db()
        try:
            query = "SELECT id, colaborador, periodo_inicio, periodo_fim, total_horas, criado_em FROM apontamentos WHERE 1=1"
            params = []

            if colaborador and colaborador.strip():
                query += " AND LOWER(colaborador) LIKE LOWER(?)"
                params.append(f"%{colaborador.strip()}%")

            if mes and mes.strip():
                # mes no formato YYYY-MM → converter para dd/mm/yyyy parcial
                try:
                    year, month = mes.strip().split("-")
                    # Usar % para casar qualquer dia (01-31, 1-9 com 2 dígitos)
                    query += " AND (periodo_inicio LIKE ? OR periodo_fim LIKE ?)"
                    params.append(f"%/{month}/{year}")
                    params.append(f"%/{month}/{year}")
                except ValueError:
                    pass  # ignorar filtro de mês inválido

            query += " ORDER BY id DESC LIMIT 100"
            rows = conn.execute(query, params).fetchall()
        finally:
            conn.close()

        result = [
            {
                "id": row["id"],
                "colaborador": row["colaborador"],
                "periodo_inicio": row["periodo_inicio"],
                "periodo_fim": row["periodo_fim"],
                "total_horas": row["total_horas"],
                "criado_em": row["criado_em"],
            }
            for row in rows
        ]
        return {"registros": result}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar histórico: {str(e)}")


@app.get("/api/historico/{record_id}")
async def get_historico_detail(record_id: int):
    """Retorna detalhe completo de um apontamento (incluindo dias e intervalos)."""
    try:
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT * FROM apontamentos WHERE id = ?", (record_id,)
            ).fetchone()
        finally:
            conn.close()

        if not row:
            raise HTTPException(status_code=404, detail="Registro não encontrado")

        return {
            "id": row["id"],
            "colaborador": row["colaborador"],
            "periodo_inicio": row["periodo_inicio"],
            "periodo_fim": row["periodo_fim"],
            "total_horas": row["total_horas"],
            "criado_em": row["criado_em"],
            "dias": json.loads(row["dados_json"]),
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar registro: {str(e)}")


@app.delete("/api/historico/{record_id}")
async def delete_historico(record_id: int):
    """Remove um apontamento do histórico pelo ID."""
    try:
        conn = get_db()
        try:
            result = conn.execute(
                "DELETE FROM apontamentos WHERE id = ?", (record_id,)
            )
            conn.commit()
            deleted = result.rowcount
        finally:
            conn.close()

        if deleted == 0:
            raise HTTPException(status_code=404, detail="Registro não encontrado")

        return {"mensagem": "Registro excluído com sucesso", "id": record_id}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao excluir registro: {str(e)}")


# Servir arquivos estáticos do frontend na raiz
frontend_dir = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_dir):
    app.mount("/", StaticFiles(directory=frontend_dir, html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)



